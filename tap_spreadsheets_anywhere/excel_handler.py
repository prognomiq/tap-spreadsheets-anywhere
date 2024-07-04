import re
import openpyxl
import logging

import xlrd
from openpyxl.cell import Cell

LOGGER = logging.getLogger(__name__)

def generator_wrapper(reader, no_key_formatting=False, header_row=None, skip_preceding=0):
    for row_index, row in enumerate(reader):
        to_return = {}
        if header_row is None:
            if row_index >= skip_preceding:
                header_row = row
            continue

        for index, cell in enumerate(row):
            header_cell = header_row[index]
            formatted_key = header_cell.value
            if not formatted_key:
                formatted_key = '' # default to empty string for key
            formatted_key = str(formatted_key)

            if not no_key_formatting:
                # replace non-word characters with underscores
                formatted_key = re.sub(r"\W+", '_', formatted_key).strip('_')
                formatted_key = formatted_key.lower()

            to_return[formatted_key] = cell.value
        yield to_return

def get_legacy_row_iterator(table_spec, file_handle):
    workbook = xlrd.open_workbook(on_demand=True,file_contents=file_handle.read())
    if "worksheet_name" in table_spec:
        try:
            sheet = workbook.sheet_by_name(table_spec["worksheet_name"])
        except Exception as e:
            LOGGER.error("Unable to open specified sheet '"+table_spec["worksheet_name"]+"' - did you check the workbook's sheet name for spaces?")
            raise e
    else:
        try:
            sheet_name_list = workbook.sheet_names()
            #if one sheet
            if(workbook.nsheets == 1):
                sheet = workbook.sheet_by_name(sheet_name_list[0])
            #else picks sheet with most data found determined by number of rows
            else:
                sheet_list = workbook.sheets()
                max_row = 0
                max_name = ""
                for i in sheet_list:
                    if i.nrows > max_row:
                        max_row = i.nrows
                        max_name = i.name
                sheet = workbook.sheet_by_name(max_name)
        except Exception as e:
            LOGGER.info(e)
            sheet = workbook.sheet_by_name(sheet_name_list[0])
    return generator_wrapper(sheet.get_rows())


def get_row_iterator(table_spec, file_handle, skip_preceding=0):
    workbook = openpyxl.load_workbook(file_handle, read_only=True)

    if "worksheet_name" in table_spec:
        try:
            active_sheet = workbook[table_spec["worksheet_name"]]
        except Exception as e:
            LOGGER.error("Unable to open specified sheet '"+table_spec["worksheet_name"]+"' - did you check the workbook's sheet name for spaces?")
            raise e
    else:
        try:
            worksheets = workbook.worksheets
            #if one sheet
            if(len(worksheets) == 1):
                active_sheet = worksheets[0]
            #else picks sheet with most data found determined by number of rows
            else:
                max_row = 0
                longest_sheet_index = 0
                for i, sheet in enumerate(worksheets):
                    if sheet.max_row > max_row:
                        max_row = i.max_row
                        longest_sheet_index = i
                active_sheet = worksheets[longest_sheet_index]
        except Exception as e:
            LOGGER.info(e)
            active_sheet = worksheets[0]

    no_key_formatting = table_spec.get('no_key_formatting', False)

    field_names = table_spec.get('field_names', None)
    if field_names:
        header_row = [Cell(active_sheet, value=f) for f in field_names]
    else:
        header_row = None

    return generator_wrapper(active_sheet, no_key_formatting, header_row, skip_preceding)
